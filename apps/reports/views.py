# apps/reports/views.py
from django.shortcuts import render
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .services import ReportService
from .models import SavedReport
from .serializers import SavedReportSerializer
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from rest_framework import status
import traceback
import logging

class ReportsView(LoginRequiredMixin, TemplateView):
    """View principal da página de relatórios"""
    template_name = 'reports/reports.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Relatórios'
        return context

class ReportsAPIView(APIView):
    """API para dados dos relatórios"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            period = request.GET.get('period', 'month')
            
            print("="*50)
            print("🔍 DEBUG REPORTS API")
            print(f"Usuário: {request.user} (ID: {request.user.id})")
            print(f"Autenticado: {request.user.is_authenticated}")
            print(f"Período: {period}")
            print(f"Method: {request.method}")
            print(f"Headers: {dict(request.headers)}")
            
            # Teste simples primeiro
            from sales.models import Sale
            total_vendas = Sale.objects.filter(status='Done').count()
            print(f"Total vendas Done no banco: {total_vendas}")
            
            if total_vendas == 0:
                print("⚠️ ATENÇÃO: Nenhuma venda com status 'Done' encontrada!")
                
                # Listar todos os status existentes
                status_list = Sale.objects.values_list('status', flat=True).distinct()
                print(f"Status existentes no banco: {list(status_list)}")
                
                # Mostrar exemplo de venda
                venda_exemplo = Sale.objects.first()
                if venda_exemplo:
                    print(f"Exemplo de venda - ID: {venda_exemplo.id}, Status: '{venda_exemplo.status}'")
            
            # Busca dados dos relatórios
            print("\n📊 Buscando vendas por marca...")
            vendas_por_marca = ReportService.get_vendas_por_marca(period)
            print(f"Resultado: {vendas_por_marca}")
            
            print("\n👥 Buscando performance vendedores...")
            performance_vendedores = ReportService.get_performance_vendedores(period)
            print(f"Resultado: {performance_vendedores}")
            
            print("\n📈 Buscando métricas gerais...")
            metricas_gerais = ReportService.get_metricas_gerais(period)
            print(f"Resultado: {metricas_gerais}")
            
            print("✅ Debug concluído, retornando resposta")
            print("="*50)
            
            return Response({
                'vendas_por_marca': vendas_por_marca,
                'performance_vendedores': performance_vendedores,
                'metricas_gerais': metricas_gerais,
                'relatorio_detalhado': ReportService.get_relatorio_detalhado(period),
                'dados_graficos': ReportService.get_dados_graficos(period),
            })
            
        except Exception as e:
            print("❌ ERRO NA API:")
            print(f"Tipo: {type(e).__name__}")
            print(f"Mensagem: {str(e)}")
            print("Traceback:")
            traceback.print_exc()
            
            return Response(
                {'error': str(e), 'type': type(e).__name__},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExportReportView(APIView):
    """Exportar relatório em diferentes formatos"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        format = request.GET.get('format', 'excel')
        period = request.GET.get('period', 'month')
        
        if format == 'excel':
            return self.export_excel(period)
        elif format == 'csv':
            return self.export_csv(period)
        elif format == 'pdf':
            return self.export_pdf(period)
        else:
            return Response({'error': 'Formato não suportado'}, status=400)
    
    def export_excel(self, period):
        """Exporta relatório para Excel"""
        # Busca dados
        dados = ReportService.get_relatorio_detalhado(period)
        metricas = ReportService.get_metricas_gerais(period)
        
        # Cria workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Vendas"
        
        # Título
        ws.merge_cells('A1:E1')
        title = ws['A1']
        title.value = f"Relatório de Vendas - {timezone.now().strftime('%d/%m/%Y')}"
        title.font = Font(size=14, bold=True)
        title.alignment = Alignment(horizontal='center')
        
        # Métricas
        ws['A3'] = "Métricas Gerais"
        ws['A3'].font = Font(bold=True)
        
        metrics_data = [
            ['Ticket Médio', f"R$ {metricas['ticket_medio']:,.2f}"],
            ['Taxa de Conversão', f"{metricas['taxa_conversao']}%"],
            ['Leads no Mês', metricas['leads_mes']],
            ['Total de Vendas', metricas['total_vendas']],
        ]
        
        for i, (label, value) in enumerate(metrics_data):
            ws[f'A{4+i}'] = label
            ws[f'B{4+i}'] = value
        
        # Cabeçalho da tabela
        headers = ['Período', 'Vendas', 'Receita', 'Comissões', 'Crescimento']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for row, item in enumerate(dados, 8):
            ws.cell(row=row, column=1).value = item['periodo']
            ws.cell(row=row, column=2).value = item['vendas']
            ws.cell(row=row, column=3).value = item['receita']
            ws.cell(row=row, column=4).value = item['comissoes']
            ws.cell(row=row, column=5).value = f"{item['crescimento']}%"
            
            # Formatação de moeda
            ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=4).number_format = 'R$ #,##0.00'
        
        # Ajusta largura das colunas
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Salva em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Retorna arquivo
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_vendas_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    def export_csv(self, period):
        """Exporta relatório para CSV"""
        dados = ReportService.get_relatorio_detalhado(period)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="relatorio_vendas_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Período', 'Vendas', 'Receita', 'Comissões', 'Crescimento'])
        
        for item in dados:
            writer.writerow([
                item['periodo'],
                item['vendas'],
                f"R$ {item['receita']:,.2f}",
                f"R$ {item['comissoes']:,.2f}",
                f"{item['crescimento']}%"
            ])
        
        return response
    
    def export_pdf(self, period):
        """Exporta relatório para PDF"""
        dados = ReportService.get_relatorio_detalhado(period)
        metricas = ReportService.get_metricas_gerais(period)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph(f"Relatório de Vendas - {timezone.now().strftime('%d/%m/%Y')}", styles['Title'])
        elements.append(title)
        
        # Espaço
        elements.append(Paragraph("<br/>", styles['Normal']))
        
        # Métricas
        elements.append(Paragraph("Métricas Gerais", styles['Heading2']))
        metrics_data = [
            ['Ticket Médio:', f"R$ {metricas['ticket_medio']:,.2f}"],
            ['Taxa de Conversão:', f"{metricas['taxa_conversao']}%"],
            ['Leads no Mês:', str(metricas['leads_mes'])],
            ['Total de Vendas:', str(metricas['total_vendas'])],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[150, 150])
        metrics_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(metrics_table)
        
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        # Tabela de dados
        elements.append(Paragraph("Vendas por Período", styles['Heading2']))
        
        table_data = [['Período', 'Vendas', 'Receita', 'Comissões', 'Crescimento']]
        for item in dados:
            table_data.append([
                item['periodo'],
                str(item['vendas']),
                f"R$ {item['receita']:,.2f}",
                f"R$ {item['comissoes']:,.2f}",
                f"{item['crescimento']}%"
            ])
        
        table = Table(table_data, colWidths=[120, 60, 100, 100, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.cyan),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_vendas_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response

class SavedReportsView(APIView):
    """Views para relatórios salvos"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        reports = SavedReport.objects.filter(created_by=request.user)
        serializer = SavedReportSerializer(reports, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        data = request.data.copy()
        data['created_by'] = request.user.id
        
        serializer = SavedReportSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)